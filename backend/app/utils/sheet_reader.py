import openpyxl
import csv
from app.schemas.event_row import EventRow

def parse_sheet(file_path):
    ext = file_path.split(".")[-1].lower()

    if ext == "csv":
        return parse_csv(file_path)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header = [str(cell.value).strip() for cell in ws[1]]
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(header, row))

        results.append(EventRow(
            slno=int(data["slno"]),
            topic=data["topic"],
            imageGenerated=str(data["imageGenerated"]).lower() == "true",
            selectDate=str(data["selectDate"]),
            time=str(data["time"])
        ))
    return results


def parse_csv(file_path):
    results = []
    with open(file_path, newline='', encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for data in reader:
            results.append(EventRow(
                slno=int(data["slno"]),
                topic=data["topic"],
                imageGenerated=str(data["imageGenerated"]).lower() == "true",
                selectDate=str(data["selectDate"]),
                time=str(data["time"])
            ))
    return results